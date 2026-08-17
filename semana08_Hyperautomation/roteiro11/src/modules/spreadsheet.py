from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def registrar_cadastro(
    dados: dict,
    caminho_planilha: Path
) -> None:
    """
    Registra um cadastro aprovado na primeira
    linha vazia da Planilha Mestra.
    """

    if not caminho_planilha.exists():
        raise FileNotFoundError(
            f"Planilha não encontrada: {caminho_planilha}"
        )

    workbook = load_workbook(caminho_planilha)
    planilha = workbook.active

    # Procura a primeira linha vazia,
    # começando depois do cabeçalho.
    linha_destino = 2

    while planilha.cell(
        row=linha_destino,
        column=1
    ).value is not None:
        linha_destino += 1

    valores = [
        dados["cpf"],
        dados["nome"],
        dados["data_nascimento"],
        dados["endereco"],
        dados["email"],
        dados["telefone"],
        "APROVADO",
        datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S"
        ),
        "",
    ]

    for coluna, valor in enumerate(
        valores,
        start=1
    ):
        planilha.cell(
            row=linha_destino,
            column=coluna,
            value=valor
        )

    workbook.save(caminho_planilha)

    print(
        f"Cadastro gravado na linha "
        f"{linha_destino}."
    )