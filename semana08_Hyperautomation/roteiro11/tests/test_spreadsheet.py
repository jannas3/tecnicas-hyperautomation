from openpyxl import Workbook, load_workbook

from src.modules.spreadsheet import registrar_cadastro


def test_registrar_cadastro(tmp_path):
    caminho_planilha = tmp_path / "Planilha_Mestra.xlsx"

    workbook = Workbook()
    planilha = workbook.active

    planilha.append(
        [
            "CPF",
            "Nome",
            "Data de Nascimento",
            "Endereço",
            "E-mail",
            "Telefone",
            "Status",
            "Data de Processamento",
            "Observações",
        ]
    )

    workbook.save(caminho_planilha)

    dados = {
        "cpf": "52998224725",
        "nome": "Maria Silva",
        "data_nascimento": "10/05/1995",
        "endereco": "Rua A, 100",
        "email": "maria@email.com",
        "telefone": "92999999999",
    }

    registrar_cadastro(
        dados=dados,
        caminho_planilha=caminho_planilha,
    )

    workbook = load_workbook(caminho_planilha)
    planilha = workbook.active

    assert planilha.max_row == 2

    assert planilha["A2"].value == "52998224725"
    assert planilha["B2"].value == "Maria Silva"
    assert planilha["G2"].value == "APROVADO"
